# _*_ coding: utf-8 _*_
# @Author  : XuTusheng
# @Time    : 2020/3/3
# @PROJECT : di5cheng-IT-PaaS
# @File    : ParseExcel.py

from openpyxl import load_workbook
from datetime import datetime
from common import config


class ParseExcel:
    """
    Use openpyxl to parse excel file data class
    """
    def __init__(self):
        # 加载excel文件到内存
        self.excelPath = config.get_excel_demo()
        self.wb = load_workbook(self.excelPath)

    def get_row_value(self, sheet_name, row_no):
        """获取某一行的数据"""
        sh = self.wb[sheet_name]
        row_value_list = []
        for y in range(2, sh.max_column + 1):
            value = sh.cell(row_no, y).value
            row_value_list.append(value)
        return row_value_list

    def get_col_value(self, sheet_name, col_no):
        """获取某一列的数据"""
        sh = self.wb[sheet_name]
        col_value_list = []
        for x in range(2, sh.max_row + 1):
            value = sh.cell(x, col_no).value
            col_value_list.append(value)
        return col_value_list

    def get_cell_of_value(self, sheet_name, row_no, col_no):
        """获取某一个单元格的数据"""
        sh = self.wb[sheet_name]
        value = sh.cell(row_no, col_no).value
        return value

    def write_cell(self, sheet_name, row_no, col_no, value):
        """向某个单元格写入数据"""
        sh = self.wb[sheet_name]
        sh.cell(row_no, col_no).value = value
        self.wb.save(self.excelPath)

    def write_current_time(self, sheet_name, row_no, col_no):
        """向某个单元格写入当前时间"""
        sh = self.wb[sheet_name]
        time = datetime.now()
        current_time = time.strftime('%Y:%m:%d %H:%M:%S')
        sh.cell(row_no, col_no).value = current_time
        self.wb.save(self.excelPath)
